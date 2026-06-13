"""
UI Benchmark Comparison router.

POST /api/benchmark
    Accepts a multipart upload of an XLSX/CSV containing baseline + current
    UI performance metrics and returns a structured comparison matrix.

POST /api/benchmark/json
    Same but accepts pre-parsed rows in JSON body.

Supports two formats:
  A) Simple flat file: Transaction | Baseline (sec) | Current (sec) [| SLA]
  B) Multi-sheet PE validation XLSX with auto-detected sheet types:
       - Sheets with "fill" + "rate" in name      → PROD vs TEST fill rates
       - Sheets with "batch" + "run/time" in name  → PROD vs TEST batch runtimes
       - Sheets with "ui" + "perf" in name          → PROD vs TEST UI load times
       - Sheets with "observation" or "sit" in name → Issue tracker
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from pydantic import BaseModel, ConfigDict

from services import config_store

router = APIRouter()


# ── Models ───────────────────────────────────────────────────────────────────

class BenchmarkRow(BaseModel):
    transaction:    str
    baseline_sec:   float
    current_sec:    float
    delta_pct:      float     # positive = regression, negative = improvement
    sla_sec:        Optional[float] = None
    sla_breach:     bool = False
    status:         str   # OK | WATCH | BREACH | N/A   (replaces GREEN/AMBER/RED)
    category:       str = ""
    pass_fail:      Optional[str] = None
    # Template-aligned extra fields
    action:         Optional[str] = None   # Load | Export | Save | Import | SRE Process Run
    records:        Optional[int] = None   # No. of Records
    concurrent_users: Optional[int] = None
    comments:       Optional[str] = None
    worksheet:      Optional[str] = None   # SRE Process / FE / Worksheet name
    # Legacy fill-rate fields
    export_time_baseline: Optional[float] = None
    export_time_current:  Optional[float] = None
    rows_baseline:  Optional[int] = None
    rows_current:   Optional[int] = None


class BenchmarkCategory(BaseModel):
    """Summary stats for one category (sheet)."""
    name:       str
    total:      int = 0
    passed:     int = 0
    failed:     int = 0
    degraded:   int = 0
    avg_delta:  float = 0.0


class BenchmarkResponse(BaseModel):
    filename:           str
    total_transactions: int
    degraded:           int      # > threshold %
    improved:           int
    unchanged:          int
    sla_breaches:       int
    avg_delta_pct:      float
    threshold_pct:      float
    rows:               List[BenchmarkRow]
    summary:            str
    categories:         List[BenchmarkCategory] = []
    fill_rate:          Optional[List[Dict[str, Any]]] = None
    observations:       Optional[List[Dict[str, Any]]] = None
    batch_perf_summary: Optional[Dict[str, Any]] = None
    evidence_sentences: List[str] = []
    coverage_summary:   Optional[Dict[str, Any]] = None
    ai_narrative:       Optional[str] = None
    ai_model:           Optional[str] = None


class JsonBenchmarkRequest(BaseModel):
    model_config = ConfigDict(extra="allow")
    rows:       List[Dict[str, Any]]
    threshold:  Optional[float] = None
    filename:   Optional[str]   = "benchmark"


# ── Column detection ─────────────────────────────────────────────────────────

_TXN_COLS   = ["transaction", "txn", "page", "scenario", "test_case", "name",
               "transaction_name", "txnname", "step", "action",
               "worksheet", "worksheet_name", "page_name", "job", "job_name"]
_BASE_COLS  = ["baseline", "baseline_sec", "baseline_time", "ref", "reference",
               "base", "expected", "base_time", "previous", "prod",
               "prod_time", "production", "production_time"]
_CUR_COLS   = ["current", "current_sec", "current_time", "test", "actual",
               "result", "cur", "latest", "new",
               "uat", "uat_time", "uat_sec", "test_time"]
_SLA_COLS   = ["sla", "sla_sec", "target", "sla_limit", "threshold", "limit",
               "max_allowed"]


def _find_col(headers: list[str], candidates: list[str]) -> str | None:
    h_map = {h.lower().replace(" ", "_").replace("-", "_"): h for h in headers}
    for c in candidates:
        if c in h_map:
            return h_map[c]
    # Fuzzy partial match
    for c in candidates:
        for h_norm, h_orig in h_map.items():
            if c in h_norm or h_norm in c:
                return h_orig
    return None


def _safe_float(v, default=0.0):
    if v is None:
        return default
    try:
        f = float(v)
        return f if f == f else default  # NaN guard
    except (ValueError, TypeError):
        return default


def _safe_int(v, default=None):
    if v is None:
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


# ── Core computation ─────────────────────────────────────────────────────────

def _compute_benchmark(rows: list[dict], threshold_pct: float) -> tuple[list[BenchmarkRow], str]:
    """Compute per-row status using the PE audit rules:
    - OK (GREEN):    Delta ≤ +threshold_pct% AND current ≤ SLA
    - WATCH (AMBER): threshold_pct < Delta ≤ threshold_pct*2 OR within 10% of SLA
    - BREACH (RED):  Delta > threshold_pct*2 OR current > SLA
    - N/A:           No baseline (baseline_sec = 0)
    """
    try:
        from services import pe_config
        action_sla_map = dict(pe_config.BENCHMARK_ACTION_SLA)
    except Exception:
        action_sla_map = {"Load": 3.0, "Export": 10.0, "Save": 5.0,
                          "Import": 15.0, "SRE Process Run": 30.0}

    results: list[BenchmarkRow] = []

    for r in rows:
        txn     = str(r.get("transaction", r.get("txn", r.get("name", "?"))))
        base    = _safe_float(r.get("baseline_sec", r.get("baseline", 0)))
        cur     = _safe_float(r.get("current_sec",  r.get("current",  0)))
        cat     = str(r.get("category", ""))
        pf      = r.get("pass_fail")
        action  = str(r.get("action", "")).strip() or None
        records = r.get("records")
        concur  = r.get("concurrent_users")
        comment = r.get("comments") or r.get("comment")
        ws_name = r.get("worksheet") or r.get("sre_process")

        # Resolve SLA: explicit column > action-type default
        sla_raw = r.get("sla_sec") or r.get("sla")
        if sla_raw:
            sla = float(sla_raw)
        elif action and action in action_sla_map and action_sla_map[action] > 0:
            sla = action_sla_map[action]
        else:
            sla = None

        watch_boundary  = threshold_pct            # e.g. 10%
        breach_boundary = threshold_pct * 2        # e.g. 20%

        if base == 0:
            delta = 0.0
            if sla and cur > sla:
                status = "BREACH"
            elif cur > 0:
                status = "N/A"
            else:
                status = "N/A"
        else:
            delta = round((cur - base) / base * 100, 2)
            sla_near = sla and cur > sla * 0.9     # within 10% of SLA
            sla_over = sla and cur > sla

            if delta > breach_boundary or sla_over:
                status = "BREACH"
            elif delta > watch_boundary or sla_near:
                status = "WATCH"
            else:
                status = "OK"

        # Override from original Pass/Fail column
        if pf and str(pf).strip().lower() == "fail" and status not in ("BREACH",):
            status = "BREACH"

        sla_breach = bool(sla and cur > sla)

        results.append(BenchmarkRow(
            transaction=txn, baseline_sec=base, current_sec=cur,
            delta_pct=delta, sla_sec=sla, sla_breach=sla_breach, status=status,
            category=cat, pass_fail=pf,
            action=action, records=int(records) if records else None,
            concurrent_users=int(concur) if concur else None,
            comments=str(comment).strip() if comment else None,
            worksheet=str(ws_name).strip() if ws_name else None,
            export_time_baseline=r.get("export_time_baseline"),
            export_time_current=r.get("export_time_current"),
            rows_baseline=r.get("rows_baseline"),
            rows_current=r.get("rows_current"),
        ))

    # Summary
    breaches = sum(1 for r in results if r.status == "BREACH")
    watches  = sum(1 for r in results if r.status == "WATCH")
    oks      = sum(1 for r in results if r.status == "OK")
    total    = len(results)

    if breaches == 0 and watches == 0:
        summary = f"✅ All {total} transactions within tolerance. No regressions detected."
    elif breaches > 0:
        summary = (f"⚠️ {breaches}/{total} transactions BREACH threshold. "
                   f"{watches} WATCH. Review red rows before go-live.")
    else:
        summary = f"⚠️ {watches} WATCH finding(s). Monitor before go-live."

    return results, summary


# ── Parse XLSX/CSV ────────────────────────────────────────────────────────────

def _is_excel(raw_bytes: bytes, filename: str) -> bool:
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    if ext in ("xlsx", "xls", "xlsm", "xlsb"):
        return True
    if len(raw_bytes) >= 4 and raw_bytes[:4] == b'PK\x03\x04':
        return True
    if len(raw_bytes) >= 4 and raw_bytes[:4] == b'\xd0\xcf\x11\xe0':
        return True
    return False


def _is_legacy_xls(raw_bytes: bytes, filename: str) -> bool:
    if len(raw_bytes) >= 4 and raw_bytes[:4] == b'\xd0\xcf\x11\xe0':
        return True
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    return ext == "xls"


# ── Sheet type detection ──────────────────────────────────────────────────────

def _classify_sheet(name: str) -> str:
    """Classify sheet by name into a PE benchmark category.

    Recognises TEST and UAT as the 'current' environment.
    Recognises env-comparison summary sheets (Env Summary, Insights, Comparison).
    """
    nl = name.lower().strip()
    if "fill" in nl and "rate" in nl:
        return "fill_rate"
    if ("batch" in nl or "runtime" in nl) and ("run" in nl or "time" in nl or "runtime" in nl):
        return "batch"
    if ("ui" in nl and "perf" in nl) or ("uat" in nl and "perf" in nl):
        return "ui_perf"
    if "observation" in nl or "sit" in nl:
        return "observations"
    # Env-comparison summary sheets
    if any(k in nl for k in ("env summary", "env_summary", "insights", "comparison")):
        return "env_comparison"
    # Sheets named "UAT <something>" or "<something> UAT" → ui_perf
    if "uat" in nl and not any(k in nl for k in ["batch", "run", "runtime"]):
        return "ui_perf"
    if any(k in nl for k in ["demand", "esp", "need", "supply"]):
        if "batch" in nl:
            return "batch"
        return "ui_perf"
    return "unknown"


def _clean_category_name(sheet_name: str, suffix: str) -> str:
    """Derive a clean category label from the sheet name.
    Strips common noise words, appends suffix if not already present."""
    name = sheet_name.strip()
    # Remove common noise suffixes
    for noise in ["run time reports", "runtime reports", "run time report",
                  "perf summary", "performance summary", "summary",
                  "report", "reports"]:
        if name.lower().endswith(noise):
            name = name[:len(name) - len(noise)].strip(" -–—_")
    # Append the suffix if not already present
    if suffix.lower() not in name.lower():
        name = f"{name} {suffix}".strip()
    return name or suffix


# ── Sheet parsers ─────────────────────────────────────────────────────────────

def _parse_ui_perf_sheet(ws, sheet_name: str) -> list[dict]:
    """Parse UI Performance Summary sheets.
    Layout: Row 1 = env labels (TEST / PROD), Row 2 = column headers,
    Row 3+ = data. Columns: Worksheet | Search | TEST Load | TEST Export | TEST Rows | (blank) | PROD Load | PROD Export | PROD Rows | Status | Comments"""
    import openpyxl
    rows_out = []
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 3 or max_c < 7:
        return []

    # Find header row: look for a row containing "loading time" or "worksheet"
    hdr_row = None
    for r in range(1, min(max_r + 1, 6)):
        vals = [str(ws.cell(r, c).value or "").lower() for c in range(1, min(max_c + 1, 12))]
        joined = " ".join(vals)
        if "loading time" in joined or "export time" in joined:
            hdr_row = r
            break

    if hdr_row is None:
        return []

    # PROD/TEST/UAT env label detection: row above headers tells us which side is which
    env_row = hdr_row - 1 if hdr_row > 1 else None
    env_labels = {}
    if env_row:
        for c in range(1, min(max_c + 1, 12)):
            v = str(ws.cell(env_row, c).value or "").strip().upper()
            if v and any(k in v for k in ["PROD", "TEST", "UAT", "BASELINE", "CURRENT"]):
                env_labels[c] = v

    # Find loading-time columns
    load_cols = []
    export_cols = []
    rows_cols = []
    status_col = None
    for c in range(1, min(max_c + 1, 15)):
        v = str(ws.cell(hdr_row, c).value or "").lower()
        if "loading" in v or "load" in v:
            load_cols.append(c)
        elif "export" in v:
            export_cols.append(c)
        elif v in ("rows", "no. of rows", "no of rows"):
            rows_cols.append(c)
        elif "status" in v:
            status_col = c

    # Also check the env row for "Status" column (some sheets put it there)
    if status_col is None and env_row:
        for c in range(1, min(max_c + 1, 15)):
            v = str(ws.cell(env_row, c).value or "").lower().strip()
            if v == "status":
                status_col = c
                break

    if len(load_cols) < 2:
        return []

    # Determine which load col is PROD vs TEST using env_labels
    test_load_col, prod_load_col = load_cols[0], load_cols[1]
    test_export_col = export_cols[0] if len(export_cols) >= 1 else None
    prod_export_col = export_cols[1] if len(export_cols) >= 2 else None
    test_rows_col = rows_cols[0] if len(rows_cols) >= 1 else None
    prod_rows_col = rows_cols[1] if len(rows_cols) >= 2 else None

    # Check if the first group is actually PROD (env label check)
    # PROD on the LEFT → swap so TEST/UAT always maps to current
    for col_idx, label in env_labels.items():
        if "PROD" in label and col_idx <= load_cols[0]:
            # PROD is on the LEFT, swap
            test_load_col, prod_load_col = prod_load_col, test_load_col
            if test_export_col and prod_export_col:
                test_export_col, prod_export_col = prod_export_col, test_export_col
            if test_rows_col and prod_rows_col:
                test_rows_col, prod_rows_col = prod_rows_col, test_rows_col
            break

    # Derive category label from actual sheet name (generic)
    cat = _clean_category_name(sheet_name, "UI Performance")

    for r in range(hdr_row + 1, max_r + 1):
        txn = str(ws.cell(r, 1).value or "").strip()
        if not txn:
            continue

        prod_load = _safe_float(ws.cell(r, prod_load_col).value)
        test_load = _safe_float(ws.cell(r, test_load_col).value)
        status_val = str(ws.cell(r, status_col).value or "").strip() if status_col else None

        rec = {
            "transaction": txn,
            "baseline_sec": prod_load,   # PROD = baseline
            "current_sec":  test_load,   # TEST = current
            "category": cat,
            "pass_fail": status_val,
        }
        if prod_export_col and test_export_col:
            rec["export_time_baseline"] = _safe_float(ws.cell(r, prod_export_col).value) or None
            rec["export_time_current"]  = _safe_float(ws.cell(r, test_export_col).value) or None
        if prod_rows_col and test_rows_col:
            rec["rows_baseline"] = _safe_int(ws.cell(r, prod_rows_col).value)
            rec["rows_current"]  = _safe_int(ws.cell(r, test_rows_col).value)

        rows_out.append(rec)

    return rows_out


def _parse_batch_sheet(ws, sheet_name: str) -> list[dict]:
    """Parse Batch Run time report sheets.
    Layout: Row with headers = Job_Name | Start | End | Status | RunTime(sec) | (blank) | Start | End | Status | RunTime(sec) | (blank) | Diff | Pass/Fail | Comments"""
    rows_out = []
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 3:
        return []

    # Find header row containing "Job_Name" or "Run Time"
    hdr_row = None
    for r in range(1, min(max_r + 1, 8)):
        vals = [str(ws.cell(r, c).value or "").lower() for c in range(1, min(max_c + 1, 16))]
        joined = " ".join(vals)
        if "job_name" in joined or ("run time" in joined and "sec" in joined):
            hdr_row = r
            break

    if hdr_row is None:
        return []

    # Find run-time columns (there should be 2 — PROD and TEST)
    rt_cols = []
    status_cols = []
    passfail_col = None
    diff_col = None
    for c in range(1, min(max_c + 1, 16)):
        v = str(ws.cell(hdr_row, c).value or "").lower().strip()
        if "run time" in v or "runtime" in v:
            rt_cols.append(c)
        elif "completion" in v or (v == "status" and not passfail_col):
            status_cols.append(c)
        elif "pass" in v and "fail" in v:
            passfail_col = c
        elif "diff" in v:
            diff_col = c

    if len(rt_cols) < 2:
        return []

    # Determine PROD vs TEST/UAT: check env row above headers
    env_row = hdr_row - 1 if hdr_row > 1 else None
    prod_rt_col, test_rt_col = rt_cols[0], rt_cols[1]
    if env_row:
        for c in range(1, min(max_c + 1, 16)):
            v = str(ws.cell(env_row, c).value or "").strip().upper()
            # If TEST or UAT label appears to the LEFT of first runtime col → swap
            if any(k in v for k in ("TEST", "UAT")) and c <= rt_cols[0]:
                prod_rt_col, test_rt_col = rt_cols[1], rt_cols[0]
                break

    cat = _clean_category_name(sheet_name, "Batch Runtime")

    for r in range(hdr_row + 1, max_r + 1):
        job = str(ws.cell(r, 1).value or "").strip()
        if not job:
            continue
        # Skip section headers (e.g. "Daily Run", "Weekly Run" sub-sections)
        prod_rt = _safe_float(ws.cell(r, prod_rt_col).value)
        test_rt = _safe_float(ws.cell(r, test_rt_col).value)
        pf = str(ws.cell(r, passfail_col).value or "").strip() if passfail_col else None

        rows_out.append({
            "transaction": job,
            "baseline_sec": prod_rt,
            "current_sec":  test_rt,
            "category": cat,
            "pass_fail": pf,
        })

    return rows_out


def _parse_fill_rate_sheet(ws) -> list[dict]:
    """Parse Fill Rate sheet — multiple date blocks, each with Type | PROD | TEST | Diff | Comments."""
    entries = []
    max_r = ws.max_row or 0
    current_date = None

    for r in range(1, max_r + 1):
        a = str(ws.cell(r, 1).value or "").strip()
        if not a:
            continue

        # Detect date row (DD.MM.YYYY or YYYY-MM-DD)
        if re.match(r"^\d{2}\.\d{2}\.\d{4}$", a) or re.match(r"^\d{4}-\d{2}-\d{2}", a):
            current_date = a
            continue

        # Skip label rows
        al = a.lower()
        if al in ("type", "live data", "live", "") or al.startswith("live"):
            continue

        # Data row
        prod_val = _safe_float(ws.cell(r, 2).value)
        test_val = _safe_float(ws.cell(r, 3).value)
        diff_val = _safe_float(ws.cell(r, 4).value)
        comment  = str(ws.cell(r, 5).value or "").strip()

        if prod_val > 0 or test_val > 0:
            entries.append({
                "date": current_date,
                "type": a,
                "prod": round(prod_val, 6),
                "test": round(test_val, 6),
                "diff": round(diff_val, 6),
                "status": comment if comment else ("Pass" if abs(diff_val) < 1.0 else "Fail"),
            })

    return entries


def _parse_observations_sheet(ws) -> list[dict]:
    """Parse SIT Observations sheet."""
    obs = []
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 2:
        return []

    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, min(max_c + 1, 10))]

    for r in range(2, max_r + 1):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, min(max_c + 1, 10))]
        if not any(vals):
            continue
        entry = {}
        for i, h in enumerate(headers):
            if i < len(vals):
                entry[h.lower().replace(" ", "_").replace("/", "_")] = vals[i]
        obs.append(entry)

    return obs


# ── Single-time capture parser (0H 4M 36S / HH:MM:SS) ────────────────────────

def _parse_hms_duration(v) -> float:
    """Parse duration strings → seconds.
    Handles: '0H 4M 36S', '4M 36S', 'HH:MM:SS', 'MM:SS', '14s', '4m 36s', '276'.
    Returns 0.0 if unparseable."""
    if v is None:
        return 0.0
    s = str(v).strip()

    # "0H 4M 36S" or "0H4M36S" — full hours/mins/secs
    m = re.match(r'(\d+)\s*[Hh]\s*(\d+)\s*[Mm]\s*(\d+(?:\.\d+)?)\s*[Ss]?', s)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))

    # "4M 36S" or "4m36s" — no hours
    m = re.match(r'(\d+)\s*[Mm]\s*(\d+(?:\.\d+)?)\s*[Ss]?$', s)
    if m:
        return int(m.group(1)) * 60 + float(m.group(2))

    # "14s" or "276s" — plain seconds with s/sec suffix
    m = re.match(r'^(\d+(?:\.\d+)?)\s*s(?:ec)?\.?$', s, re.I)
    if m:
        return float(m.group(1))

    # "4m" — plain minutes with m suffix only
    m = re.match(r'^(\d+(?:\.\d+)?)\s*m(?:in)?\.?$', s, re.I)
    if m:
        return float(m.group(1)) * 60

    # "HH:MM:SS" or "MM:SS"
    m = re.match(r'^(\d+):(\d+):(\d+(?:\.\d+)?)$', s)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))
    m = re.match(r'^(\d+):(\d+(?:\.\d+)?)$', s)
    if m:
        return int(m.group(1)) * 60 + float(m.group(2))

    return _safe_float(v)


def _parse_two_row_header_sheet(ws, sheet_name: str) -> list[dict] | None:
    """Parse Evidence-Analysis-style sheets with two header rows.

    Row 1: section labels — 'Evidence Analysis Summary', 'Attempt #1', 'AVERAGE', 'MEDIAN' …
    Row 2: version sub-labels — 'V2019', 'V2022', 'V2019', 'V2022' …
    Row 3+: index | description | values…

    Strategy: find the AVERAGE section in row 1, pick V_old (baseline) and V_new (current)
    from row 2 at the matching columns. Falls back to first Attempt if no AVERAGE.
    """
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 3 or max_c < 5:
        return None

    # Detect two-row header: row 1 has section labels, row 2 has version codes
    r1 = [str(ws.cell(1, c).value or "").strip().lower() for c in range(1, min(max_c + 1, 25))]
    r2 = [str(ws.cell(2, c).value or "").strip().lower() for c in range(1, min(max_c + 1, 25))]

    # Row 2 must have repeated version-like patterns (v20xx, v1, v2, etc.)
    ver_pattern = re.compile(r'^v\d+$')
    ver_count = sum(1 for v in r2 if ver_pattern.match(v) or re.match(r'^v\d{4}$', v))
    if ver_count < 2:
        return None

    # Name column: first col in r1 area that has descriptive content in data rows
    name_col = None
    for c in range(1, min(max_c + 1, 5)):
        for r in range(3, min(max_r + 1, 8)):
            v = ws.cell(r, c).value
            if v and str(v).strip() and not str(v).strip().isdigit():
                name_col = c
                break
        if name_col:
            break
    if name_col is None:
        return None

    # Find "average" or "avg" in row 1 (prefer it) or fall back to first attempt group
    target_start = None
    for i, v in enumerate(r1):
        if v in ("average", "avg"):
            target_start = i  # 0-indexed → col = i+1
            break
    if target_start is None:
        # Fall back: first "attempt" group
        for i, v in enumerate(r1):
            if "attempt" in v:
                target_start = i
                break
    if target_start is None:
        return None

    # From target_start: find two adjacent version columns in row 2
    base_col = cur_col = None
    found_versions = []
    for offset in range(0, 5):
        idx = target_start + offset
        if idx >= len(r2):
            break
        if ver_pattern.match(r2[idx]) or re.match(r'^v\d{4}$', r2[idx]):
            found_versions.append((idx + 1, r2[idx]))  # (1-indexed col, version label)
    if len(found_versions) < 2:
        # Only one version — use as current only (UAT capture)
        if found_versions:
            cur_col = found_versions[0][0]
        else:
            return None
    else:
        # Sort versions to identify older (baseline) and newer (current)
        try:
            found_versions.sort(key=lambda x: int(re.sub(r'\D', '', x[1])))
        except Exception:
            pass
        base_col = found_versions[0][0]
        cur_col = found_versions[1][0]

    rows = []
    for r in range(3, max_r + 1):
        name_v = ws.cell(r, name_col).value
        cur_v = ws.cell(r, cur_col).value
        if not name_v or not cur_v:
            continue
        name_s = str(name_v).strip()
        # Skip index numbers and header-like rows
        if not name_s or name_s.isdigit() or name_s.lower() in ("none", "n/a"):
            continue
        cur_sec = _safe_float(cur_v)
        if cur_sec <= 0:
            continue
        base_sec = _safe_float(ws.cell(r, base_col).value) if base_col else 0.0
        rows.append({
            "transaction": name_s,
            "baseline_sec": base_sec,
            "current_sec": cur_sec,
            "sheet": sheet_name,
        })

    return rows if rows else None


def _parse_multi_user_avg_sheet(ws, sheet_name: str) -> list[dict] | None:
    """Parse sheets with multi-user run columns and an Average column.

    Expected pattern (Project Odyssey / similar):
        Row 4: | Sr.# | Module | UI Name | Search Used | Scenario | Data Volume | Average | User 1 | User 2 | ...
        Rows 5+: data rows

    Uses 'Average' column as the current runtime. No baseline → UAT-only capture.
    Returns None if pattern not found.
    """
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 3 or max_c < 4:
        return None

    hdr_row = None
    avg_col = None
    name_col = None
    user_cols = []

    for r in range(1, min(max_r + 1, 10)):
        row_vals = [str(ws.cell(r, c).value or "").lower().strip() for c in range(1, min(max_c + 1, 25))]
        # Detect "average" + "user N" pattern
        avg_indices = [i for i, v in enumerate(row_vals) if v in ("average", "avg")]
        user_indices = [i for i, v in enumerate(row_vals) if re.match(r'^user\s*\d+$', v)]
        if avg_indices and (user_indices or any("ui name" in v or "ui_name" in v for v in row_vals)):
            hdr_row = r
            avg_col = avg_indices[0] + 1
            user_cols = [i + 1 for i in user_indices]
            # Name column: prefer "ui name", fall back to first text-ish col
            for i, v in enumerate(row_vals):
                if any(k in v for k in ("ui name", "ui_name", "transaction", "process", "name", "scenario")):
                    name_col = i + 1
                    break
            break

    if hdr_row is None or avg_col is None or name_col is None:
        return None

    rows = []
    for r in range(hdr_row + 1, max_r + 1):
        name_v = ws.cell(r, name_col).value
        avg_v = ws.cell(r, avg_col).value
        if not name_v or not avg_v:
            continue
        name_s = str(name_v).strip()
        if not name_s or name_s.lower() in ("", "none", "n/a", "ui name", "name"):
            continue

        # Average column — try numeric first, then duration string
        avg_sec = _safe_float(avg_v)
        if avg_sec == 0.0:
            avg_sec = _parse_hms_duration(avg_v)
        if avg_sec <= 0:
            continue

        rows.append({
            "transaction": name_s,
            "baseline_sec": 0.0,
            "current_sec": avg_sec,
            "sheet": sheet_name,
        })

    return rows if rows else None



def _parse_single_time_sheet(ws, sheet_name: str) -> list[dict] | None:
    """Parse a single-environment UAT performance capture sheet.

    Expected format (Go-Live benchmark template):
        SRE Process/FE/Worksheet | Search | Records | Action | TimeTaken | Users | Comments

    There is NO baseline column — just the UAT/current time.
    baseline_sec is set to 0 → _compute_benchmark will mark these as 'N/A' status.
    Transaction label = '{Worksheet} — {Action}' for uniqueness.

    Detection criteria:
    - Has a column whose name contains 'time' (but not 'start' or 'end')
    - Values in that column match XH YM ZS or HH:MM:SS patterns
    - Has a worksheet/process/transaction name column
    """
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 2 or max_c < 2:
        return None

    # Find header row (first 5 rows)
    hdr_row = None
    for r in range(1, min(max_r + 1, 6)):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, min(max_c + 1, 10))]
        if any(v for v in vals):
            hdr_row = r
            break
    if hdr_row is None:
        return None

    headers = [str(ws.cell(hdr_row, c).value or "").strip() for c in range(1, min(max_c + 1, 12))]
    hn = [h.lower().replace(" ", "_").replace("/", "_").replace(".", "_") for h in headers]

    # Time column: contains 'time' but not 'start'/'end'/'datetime'
    time_col = None
    for i, h in enumerate(hn):
        if "time" in h and not any(k in h for k in ("start", "end", "datetime", "date")):
            time_col = i + 1
            break
    if time_col is None:
        return None

    # Name/worksheet column — use a priority list so "job_name" beats "folder_name"
    name_col = 1
    _name_priority = [
        ("job_name",),                          # exact: job_name col
        ("process",),                           # process/sre column
        ("transaction",),                       # UI transaction col
        ("worksheet", "sre"),                   # worksheet name
        ("task",),
        ("step",),
        ("job",),                               # generic job col (not job_id)
        ("name",),                              # last resort — avoid matching folder_name
    ]
    for keywords in _name_priority:
        found = False
        for i, h in enumerate(hn):
            if h in ("job_id", "jobid", "id", "run_id", "folder_name", "folder"):
                continue
            # For generic "name" — only match if it IS the header (not contains)
            if "name" in keywords:
                if h == "name" or h.endswith("_name") and not any(
                        skip in h for skip in ("folder", "application", "status")):
                    name_col = i + 1
                    found = True
                    break
            elif any(k in h for k in keywords):
                name_col = i + 1
                found = True
                break
        if found:
            break

    # Action column (optional)
    action_col = None
    for i, h in enumerate(hn):
        if "action" in h:
            action_col = i + 1
            break

    # Records column (optional — "No. of Records", "records", "volume")
    records_col = None
    for i, h in enumerate(hn):
        if any(k in h for k in ("record", "volume", "no_of", "count", "qty")):
            records_col = i + 1
            break

    # Concurrent users column (optional)
    concur_col = None
    for i, h in enumerate(hn):
        if any(k in h for k in ("concurrent", "user", "users")):
            concur_col = i + 1
            break

    # Comments column (optional)
    comment_col = None
    for i, h in enumerate(hn):
        if "comment" in h or "remark" in h or "note" in h:
            comment_col = i + 1
            break

    # Verify the time column actually has parseable values.
    # Accept: XH YM ZS | HH:MM:SS | plain integers when column explicitly implies runtime.
    time_hdr_norm = hn[time_col - 1]
    col_is_runtime = any(k in time_hdr_norm for k in ("runtime", "run_time", "elapsed",
                                                       "duration", "timetaken", "time_taken"))
    valid_count = 0
    for r in range(hdr_row + 1, min(max_r + 1, hdr_row + 10)):
        v = ws.cell(r, time_col).value
        if not v:
            continue
        s = str(v).strip()
        # XH YM ZS  |  HH:MM:SS  |  14s  |  4m  |  numeric when col is runtime
        if (re.search(r'\d+[HhMmSs]|\d+:\d+', s)
                or (col_is_runtime and re.match(r'^\d+(\.\d+)?$', s) and float(s) > 0)):
            valid_count += 1
    if valid_count == 0:
        return None

    cat = _clean_category_name(sheet_name, "UAT Performance")
    rows_out = []

    for r in range(hdr_row + 1, max_r + 1):
        name = str(ws.cell(r, name_col).value or "").strip()
        if not name:
            continue
        time_val = ws.cell(r, time_col).value
        if not time_val:
            continue
        secs = _parse_hms_duration(time_val)
        if secs == 0.0:
            continue

        # Build unique transaction label: name + action if present
        action_val = None
        if action_col:
            action_val = str(ws.cell(r, action_col).value or "").strip() or None
            label = f"{name} — {action_val}" if action_val else name
        else:
            label = name

        records_val = None
        if records_col:
            rv = ws.cell(r, records_col).value
            try:
                records_val = int(float(str(rv).replace(",", ""))) if rv else None
            except Exception:
                pass

        concur_val = None
        if concur_col:
            cv = ws.cell(r, concur_col).value
            try:
                concur_val = int(float(str(cv))) if cv else None
            except Exception:
                pass

        comment_val = None
        if comment_col:
            cmt = ws.cell(r, comment_col).value
            comment_val = str(cmt).strip() if cmt else None

        rows_out.append({
            "transaction":    label,
            "baseline_sec":   0.0,
            "current_sec":    round(secs, 2),
            "category":       cat,
            "action":         action_val,
            "records":        records_val,
            "concurrent_users": concur_val,
            "comments":       comment_val,
            "worksheet":      name,   # raw worksheet/SRE process name (without action)
        })

    return rows_out if rows_out else None



def _parse_interval_runtime(v) -> float:
    """Parse runtime in Oracle interval format '+00 HH:MM:SS.ffffff' → seconds.
    Falls back to _safe_float for plain numeric values."""
    if v is None:
        return 0.0
    s = str(v).strip()
    m = re.match(r'[+-]?(\d+)\s+(\d+):(\d+):(\d+(?:\.\d+)?)', s)
    if m:
        days  = int(m.group(1))
        hours = int(m.group(2))
        mins  = int(m.group(3))
        secs  = float(m.group(4))
        return days * 86400 + hours * 3600 + mins * 60 + secs
    m = re.match(r'(\d+):(\d+):(\d+(?:\.\d+)?)', s)
    if m:
        return int(m.group(1)) * 3600 + int(m.group(2)) * 60 + float(m.group(3))
    return _safe_float(v)


# ── Generic env-comparison sheet parser ──────────────────────────────────────

def _parse_env_comparison_sheet(ws, sheet_name: str) -> list[dict] | None:
    """Parse any sheet that has a PROD (baseline) column and a TEST/UAT (current) column.

    Handles:
    - Multi-line headers: "PROD 2019\\nAvg (s)" → joined as "PROD 2019 Avg (s)"
    - Simple headers: "PROD AVG", "TEST AVG", "UAT Time"
    - Unit detection: "(min)" in header name → multiply by 60 to convert to seconds

    Returns list of row dicts or None if the PROD/TEST pattern is not found.
    """
    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 2 or max_c < 3:
        return None

    # Scan first 7 rows for a header row that contains both a PROD-like and TEST/UAT-like col
    for hdr_row in range(1, min(max_r + 1, 8)):
        # Normalize multi-line headers by joining all lines
        headers_raw = []
        for c in range(1, min(max_c + 1, 20)):
            v = ws.cell(hdr_row, c).value
            if v is None:
                headers_raw.append("")
            else:
                lines = [l.strip() for l in str(v).split("\n") if l.strip()]
                headers_raw.append(" ".join(lines))

        headers_norm = [
            h.lower().replace(" ", "_").replace("-", "_").replace("(", "").replace(")", "")
            for h in headers_raw
        ]

        # Job/name column: first col whose name looks like a job identifier
        job_col = None
        for i, hn in enumerate(headers_norm):
            if any(k in hn for k in ("job_name", "job", "name", "process",
                                     "transaction", "worksheet", "step", "task")):
                job_col = i + 1
                break
        if job_col is None:
            continue

        # PROD / BASELINE column: contains prod/baseline/base keywords OR version label (V2019, v1)
        prod_col = None
        for i, hn in enumerate(headers_norm):
            if i + 1 == job_col or not hn:
                continue
            if any(k in hn for k in ("prod", "baseline", "base_", "before",
                                     "previous", "old_", "ref_", "reference")):
                prod_col = i + 1
                break
            # Version labels: v2019, v2022, v1, v2, ver_1 etc. — pick the LOWER version as baseline
            if re.search(r'v\d{2,4}', hn) or re.match(r'v\d+$', hn) or "ver" in hn:
                prod_col = i + 1
                break
        if prod_col is None:
            continue

        # TEST / UAT / CURRENT column: comes after prod (or independently)
        # Also handles: newer version label (v2022 after v2019)
        test_col = None
        for i, hn in enumerate(headers_norm):
            if i + 1 in (job_col, prod_col) or not hn:
                continue
            if any(k in hn for k in ("test", "uat", "current", "new_", "after", "latest")):
                test_col = i + 1
                break
            # Version label that is different from prod (assume first=baseline, second=current)
            if (re.search(r'v\d{2,4}', hn) or re.match(r'v\d+$', hn) or "ver" in hn):
                test_col = i + 1
                break
        if test_col is None:
            continue

        # Verify at least one numeric value exists in prod/test cols
        has_data = False
        for r in range(hdr_row + 1, min(max_r + 1, hdr_row + 6)):
            pv = _safe_float(ws.cell(r, prod_col).value)
            tv = _safe_float(ws.cell(r, test_col).value)
            if pv > 0 or tv > 0:
                has_data = True
                break
        if not has_data:
            continue

        # Unit detection: check prod/test headers AND any other header for "(min)" hint
        prod_hdr = headers_norm[prod_col - 1]
        test_hdr = headers_norm[test_col - 1]
        all_hdrs_str = " ".join(headers_norm)
        # If any column name references minutes (but not seconds) → values are in minutes
        has_min_hint = ("min" in all_hdrs_str and
                        not any(k in all_hdrs_str for k in ("_sec", "second", "_s_", "baseline_sec")))
        # Double-check: if the prod/test values are small (<60) and no explicit sec marker → minutes
        sample_prod = _safe_float(ws.cell(hdr_row + 1, prod_col).value)
        is_minutes = (
            has_min_hint or
            (sample_prod > 0 and sample_prod < 60 and
             not any(k in prod_hdr for k in ("sec", "_s", "second")))
        )
        scale = 60.0 if is_minutes else 1.0

        cat = _clean_category_name(sheet_name, "Batch Performance")
        rows_out = []
        for r in range(hdr_row + 1, max_r + 1):
            job = str(ws.cell(r, job_col).value or "").strip()
            if not job:
                continue
            # Skip section label rows (e.g. "SECTION A —")
            if job.upper().startswith("SECTION"):
                continue
            pv = _safe_float(ws.cell(r, prod_col).value) * scale
            tv = _safe_float(ws.cell(r, test_col).value) * scale
            if pv == 0 and tv == 0:
                continue
            rows_out.append({
                "transaction": job,
                "baseline_sec": round(pv, 3),
                "current_sec":  round(tv, 3),
                "category": cat,
            })

        if rows_out:
            return rows_out

    return None


# ── Sibling-sheet join (PROD + TEST as separate sheets) ───────────────────────

def _parse_sibling_env_sheets(wb) -> list[dict] | None:
    """Detect PROD and TEST/UAT sheet pairs in the same workbook and join by job name.

    Handles workbooks where PROD and TEST runs are on separate sheets, e.g.:
        PROD2025 | TEST2025 | PROD2019
    Each sheet must have a process/job column + a runtime column.
    Oracle interval format (+DD HH:MM:SS.ffffff) is parsed automatically.
    Uses the most recent PROD sheet and most recent TEST sheet (last alphabetically).
    Returns comparison rows or None.
    """
    prod_sheets = [
        sn for sn in wb.sheetnames
        if re.search(r'prod', sn, re.I) and not re.search(r'test|uat', sn, re.I)
    ]
    test_sheets = [
        sn for sn in wb.sheetnames
        if re.search(r'test|uat', sn, re.I) and not re.search(r'prod', sn, re.I)
    ]

    if not prod_sheets or not test_sheets:
        return None

    def _extract_runs(ws) -> dict[str, list[float]]:
        """Extract {process_name: [runtime_seconds, ...]} from a single-env sheet."""
        max_r = ws.max_row or 0
        max_c = ws.max_column or 0
        if max_r < 2:
            return {}
        headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, min(max_c + 1, 15))]
        hn = [h.lower().replace(" ", "_").replace("-", "_") for h in headers]

        # Job column: prefer PROCESS/JOB_NAME over bare JOB_ID (which is numeric)
        job_col = None
        # Priority order: process → job_name → name → task → job (not job_id alone)
        for priority in (("process",), ("job_name",), ("name",), ("task",),
                         ("job",), ("step",)):
            for i, h in enumerate(hn):
                # Skip columns that are clearly just an ID number
                if h in ("job_id", "jobid", "id", "run_id", "process_id"):
                    continue
                if any(k in h for k in priority):
                    job_col = i + 1
                    break
            if job_col:
                break
        if job_col is None:
            return {}

        # Runtime column: runtime, run_time, elapsed, duration
        rt_col = next(
            (i + 1 for i, h in enumerate(hn)
             if any(k in h for k in ("runtime", "run_time", "elapsed", "duration"))),
            None
        )
        if rt_col is None:
            return {}

        runs: dict[str, list[float]] = {}
        for r in range(2, max_r + 1):
            job = str(ws.cell(r, job_col).value or "").strip()
            if not job:
                continue
            rt = _parse_interval_runtime(ws.cell(r, rt_col).value)
            if rt > 0:
                runs.setdefault(job, []).append(rt)
        return runs

    # Use the last (most recent) PROD and TEST sheet by name
    prod_runs = _extract_runs(wb[prod_sheets[-1]])
    test_runs = _extract_runs(wb[test_sheets[-1]])

    if not prod_runs or not test_runs:
        return None

    common = sorted(set(prod_runs) & set(test_runs))
    if not common:
        return None

    rows_out = []
    for job in common:
        p_avg = sum(prod_runs[job]) / len(prod_runs[job])
        t_avg = sum(test_runs[job]) / len(test_runs[job])
        rows_out.append({
            "transaction":  job,
            "baseline_sec": round(p_avg, 2),
            "current_sec":  round(t_avg, 2),
            "category":     "Batch Performance",
        })
    return rows_out if rows_out else None



# Job-column synonyms — any of these as the first header = job name column
_BP_JOB_COLS = {"job", "job_name", "job name", "task", "task_name", "batch_job",
                "step", "process", "name"}

# Runtime column keyword fragments — column names containing any of these
# AND at least two present = before/after runtime columns
_BP_RT_FRAGMENTS = ("runtime", "run_time", "run time", "elapsed", "duration",
                    "time_new", "time_old", "new_time", "old_time",
                    "time_prod", "time_test", "time_pre", "time_post",
                    "before", "after", "prod_time", "test_time",
                    "time_uat", "uat_time", "uat_rt",              # UAT environment
                    "runtime_new", "runtime_old", "runtime_prod", "runtime_test", "runtime_uat")


def _detect_batch_perf_headers(ws) -> tuple | None:
    """Return (hdr_row, first_rt_col, second_rt_col) if the sheet looks like a
    before/after batch runtime comparison, else None.

    Detection is purely column-name based — no customer-specific strings.
    Scans the first 8 rows to handle files with date/run-label rows above headers.

    Accepted job-column names:  job, job_name, task, name, step, process, …
    Accepted runtime columns:   any pair containing 'runtime', 'elapsed',
                                'duration', 'before', 'after', 'time_new', etc.
    """
    max_r = ws.max_row or 0
    for r in range(1, min(max_r + 1, 9)):
        vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, 15)]
        first = vals[0].lower().replace(" ", "_").replace("-", "_")
        if first not in _BP_JOB_COLS:
            continue
        # Collect columns whose names contain any runtime fragment
        rt_cols = [
            i + 1
            for i, v in enumerate(vals)
            if v and any(frag in v.lower().replace(" ", "_").replace("-", "_")
                         for frag in _BP_RT_FRAGMENTS)
        ]
        if len(rt_cols) >= 2:
            return r, rt_cols[0], rt_cols[1]   # (hdr_row, new_col, old_col)
    return None


def _parse_batch_perf_sheet(ws, sheet_name: str) -> list[dict] | None:
    """Parse a RUNTIME_<new>/RUNTIME_<old> batch comparison sheet.
    Returns a list of row dicts or None if the format is not detected."""
    detected = _detect_batch_perf_headers(ws)
    if detected is None:
        return None
    hdr_row, new_col, old_col = detected
    rows_out: list[dict] = []
    for r in range(hdr_row + 1, (ws.max_row or 0) + 1):
        job = str(ws.cell(r, 1).value or "").strip()
        if not job:
            continue
        raw_new = ws.cell(r, new_col).value
        raw_old = ws.cell(r, old_col).value
        new_secs = _safe_float(raw_new) if raw_new is not None else 0.0
        old_secs = _safe_float(raw_old) if raw_old is not None else 0.0
        if new_secs == 0.0 and old_secs == 0.0:
            continue  # no data for this job
        rows_out.append({
            "transaction": job,
            "baseline_sec": old_secs,   # old runtime = baseline
            "current_sec":  new_secs,   # new runtime = current
            "category": "Batch Performance",
        })
    return rows_out if rows_out else None


def _build_batch_perf_summary(rows: list[dict], threshold_pct: float) -> dict:
    """Compute regression / improvement breakdown and top-10 lists from batch perf rows."""
    regressions: list[dict] = []
    improvements: list[dict] = []
    no_change = 0
    new_only  = 0

    for r in rows:
        old, new = r["baseline_sec"], r["current_sec"]
        job = r["transaction"]
        if old == 0:
            new_only += 1
            continue
        delta_pct  = (new - old) / old * 100
        delta_secs = old - new  # positive = time saved
        entry = {
            "job":       job,
            "old_secs":  round(old,  1),
            "new_secs":  round(new,  1),
            "delta_secs": round(delta_secs, 1),
            "delta_pct":  round(delta_pct,  1),
        }
        if delta_pct > threshold_pct:
            regressions.append(entry)
        elif delta_pct < -5:
            improvements.append(entry)
        else:
            no_change += 1

    regressions.sort(key=lambda x: x["delta_pct"], reverse=True)   # worst first
    improvements.sort(key=lambda x: x["delta_pct"])                 # best first

    comparable = [r for r in rows if r["baseline_sec"] > 0]
    net_delta  = sum(r["baseline_sec"] - r["current_sec"] for r in comparable)

    return {
        "total_jobs":      len(rows),
        "comparable":      len(comparable),
        "regressions":     len(regressions),
        "improvements":    len(improvements),
        "new_only":        new_only,
        "no_change":       no_change,
        "net_delta_secs":  round(net_delta, 1),
        "top_regressions": regressions[:10],
        "top_improvements": improvements[:10],
    }


# ── Master multi-sheet parser ────────────────────────────────────────────────

def _parse_benchmark_file(raw_bytes: bytes, filename: str) -> dict:
    """Parse benchmark file. Returns a dict with keys:
       rows, fill_rate, observations, is_multi_sheet"""
    import openpyxl

    result = {"rows": [], "fill_rate": None, "observations": None,
               "is_multi_sheet": False, "batch_perf_all_rows": None,
               "debug_info": ""}

    if not _is_excel(raw_bytes, filename):
        # CSV fallback — simple flat format
        result["rows"] = _parse_flat_csv(raw_bytes, filename)
        return result

    try:
        wb = openpyxl.load_workbook(io.BytesIO(raw_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Excel parse failed: {e}") from e

    sheet_names = wb.sheetnames
    result["debug_info"] = f"Sheets found: {list(sheet_names)}."

    # ── 1. Batch performance format detection (RUNTIME_<new>/RUNTIME_<old>) ────
    for sn in sheet_names:
        bp_rows = _parse_batch_perf_sheet(wb[sn], sn)
        if bp_rows is not None:
            result["batch_perf_all_rows"] = bp_rows
            result["rows"] = bp_rows
            wb.close()
            return result

    # ── 2. Sibling-sheet env join (PROD2025 + TEST2025 separate sheets) ─────────
    sibling_rows = _parse_sibling_env_sheets(wb)
    if sibling_rows:
        result["batch_perf_all_rows"] = sibling_rows
        result["rows"] = sibling_rows
        wb.close()
        return result
    # ──────────────────────────────────────────────────────────────────────────

    if len(sheet_names) <= 1:
        ws = wb[sheet_names[0]]
        rows = _try_parse_single_sheet(ws, sheet_names[0])
        result["rows"] = rows
        wb.close()
        return result

    # Multi-sheet — classify each sheet
    result["is_multi_sheet"] = True
    all_rows = []
    sheet_debug = []

    for sn in sheet_names:
        ws = wb[sn]
        stype = _classify_sheet(sn)
        sheet_debug.append(f"{sn!r}→{stype}")

        if stype == "fill_rate":
            result["fill_rate"] = _parse_fill_rate_sheet(ws)
        elif stype == "batch":
            rows = _parse_batch_sheet(ws, sn)
            if rows:
                all_rows.extend(rows)
            else:
                # Fallthrough: might be a single-time or env-comparison sheet
                rows = (_parse_multi_user_avg_sheet(ws, sn)
                        or _parse_two_row_header_sheet(ws, sn)
                        or _parse_single_time_sheet(ws, sn)
                        or _parse_env_comparison_sheet(ws, sn)
                        or [])
                all_rows.extend(rows)
        elif stype == "ui_perf":
            rows = _parse_ui_perf_sheet(ws, sn)
            if rows:
                all_rows.extend(rows)
            else:
                # Sheet name says UI perf but layout is different — try multi-user/single-time capture
                rows = (_parse_multi_user_avg_sheet(ws, sn)
                        or _parse_two_row_header_sheet(ws, sn)
                        or _parse_single_time_sheet(ws, sn)
                        or _parse_env_comparison_sheet(ws, sn)
                        or [])
                all_rows.extend(rows)
        elif stype == "observations":
            result["observations"] = _parse_observations_sheet(ws)
        elif stype == "env_comparison":
            rows = _parse_env_comparison_sheet(ws, sn)
            if rows:
                all_rows.extend(rows)
        else:
            # Unknown — cascade through all parsers
            rows = (_parse_env_comparison_sheet(ws, sn)
                    or _parse_two_row_header_sheet(ws, sn)
                    or _parse_multi_user_avg_sheet(ws, sn)
                    or _parse_single_time_sheet(ws, sn)
                    or _try_parse_single_sheet(ws, sn)
                    or [])
            all_rows.extend(rows)

    result["rows"] = all_rows
    result["debug_info"] = f"Sheets classified: {'; '.join(sheet_debug)}. Rows extracted: {len(all_rows)}."
    wb.close()
    return result


def _try_parse_single_sheet(ws, sheet_name: str) -> list[dict]:
    """Try to parse a single sheet as env-comparison, UI perf, batch, single-time, or flat format."""
    # Env comparison (PROD vs TEST/UAT column pair)
    rows = _parse_env_comparison_sheet(ws, sheet_name)
    if rows:
        return rows
    # Two-row header with version labels (Evidence Analysis style)
    rows = _parse_two_row_header_sheet(ws, sheet_name)
    if rows:
        return rows
    # Multi-user average capture (UI Name | Average | User 1 | User 2 | ...)
    rows = _parse_multi_user_avg_sheet(ws, sheet_name)
    if rows:
        return rows
    # UI perf sheet (loading time layout)
    rows = _parse_ui_perf_sheet(ws, sheet_name)
    if rows:
        return rows
    # Batch runtime sheet
    rows = _parse_batch_sheet(ws, sheet_name)
    if rows:
        return rows
    # Single-time capture (0H 4M 36S / 14s format — UAT performance template)
    rows = _parse_single_time_sheet(ws, sheet_name)
    if rows:
        return rows
    # Flat format — only use when columns are semantically identifiable
    return _parse_flat_sheet(ws)


def _parse_flat_sheet(ws) -> list[dict]:
    """Parse a simple flat sheet: Transaction | Baseline | Current [| SLA]."""
    import pandas as pd

    max_r = ws.max_row or 0
    max_c = ws.max_column or 0
    if max_r < 2 or max_c < 3:
        return []

    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, max_c + 1)]
    txn_col  = _find_col(headers, _TXN_COLS)
    base_col = _find_col(headers, _BASE_COLS)
    cur_col  = _find_col(headers, _CUR_COLS)
    sla_col  = _find_col(headers, _SLA_COLS)

    if not txn_col:
        txn_col = headers[0] if headers else None
    if not base_col and len(headers) >= 2:
        base_col = headers[1]
    if not cur_col and len(headers) >= 3:
        cur_col = headers[2]

    if not txn_col or not base_col or not cur_col:
        return []

    h_idx = {h: i + 1 for i, h in enumerate(headers)}
    rows = []
    for r in range(2, max_r + 1):
        txn = str(ws.cell(r, h_idx.get(txn_col, 1)).value or "").strip()
        if not txn:
            continue
        base = _safe_float(ws.cell(r, h_idx.get(base_col, 2)).value)
        cur  = _safe_float(ws.cell(r, h_idx.get(cur_col, 3)).value)
        sla  = _safe_float(ws.cell(r, h_idx.get(sla_col, 0)).value) if sla_col and sla_col in h_idx else None

        rows.append({
            "transaction": txn,
            "baseline_sec": base,
            "current_sec":  cur,
            "sla_sec":      sla if sla else None,
        })
    return rows


def _parse_flat_csv(raw_bytes: bytes, filename: str) -> list[dict]:
    """Parse CSV fallback."""
    import pandas as pd
    try:
        df = pd.read_csv(io.BytesIO(raw_bytes))
    except Exception:
        try:
            df = pd.read_excel(io.BytesIO(raw_bytes), engine="openpyxl")
        except Exception as e:
            raise ValueError(f"Cannot decode file: {e}") from e

    headers = list(df.columns)
    txn_col  = _find_col(headers, _TXN_COLS)
    base_col = _find_col(headers, _BASE_COLS)
    cur_col  = _find_col(headers, _CUR_COLS)
    sla_col  = _find_col(headers, _SLA_COLS)
    if not txn_col or not base_col or not cur_col:
        cols = list(df.columns)
        if len(cols) >= 3:
            txn_col = txn_col or cols[0]
            base_col = base_col or cols[1]
            cur_col = cur_col or cols[2]
        else:
            return []

    rows = []
    for _, row in df.iterrows():
        txn = str(row.get(txn_col, "?"))
        if not txn or txn == "nan":
            continue
        base = _safe_float(row.get(base_col, 0))
        cur  = _safe_float(row.get(cur_col, 0))
        sla  = _safe_float(row.get(sla_col)) if sla_col else None
        rows.append({
            "transaction": txn,
            "baseline_sec": base,
            "current_sec": cur,
            "sla_sec": sla if sla else None,
        })
    return rows


# ── Build category summaries ──────────────────────────────────────────────────

def _build_categories(result_rows: list[BenchmarkRow]) -> list[BenchmarkCategory]:
    cats: dict[str, BenchmarkCategory] = {}
    for r in result_rows:
        c = r.category or "General"
        if c not in cats:
            cats[c] = BenchmarkCategory(name=c)
        cats[c].total += 1
        if r.status == "BREACH":
            cats[c].degraded += 1
            cats[c].failed += 1
        elif r.status == "WATCH":
            cats[c].degraded += 1
        elif r.pass_fail and r.pass_fail.strip().lower() == "fail":
            cats[c].failed += 1
        else:
            cats[c].passed += 1
    for cat in cats.values():
        vals = [r.delta_pct for r in result_rows if (r.category or "General") == cat.name]
        cat.avg_delta = round(sum(vals) / len(vals), 2) if vals else 0.0
    return list(cats.values())


def _build_coverage_summary(rows: list[BenchmarkRow]) -> dict:
    """Build coverage metadata for Band C: flows, actions, record volumes, concurrency."""
    import statistics as _stat

    flows   = sorted({(r.worksheet or r.transaction.split(" — ")[0]).strip() for r in rows if r.transaction})
    actions = sorted({r.action for r in rows if r.action})
    recs    = [r.records for r in rows if r.records and r.records > 0]
    concurs = [r.concurrent_users for r in rows if r.concurrent_users and r.concurrent_users > 0]

    return {
        "flows":         flows,
        "actions":       actions,
        "record_min":    min(recs)                       if recs else None,
        "record_max":    max(recs)                       if recs else None,
        "record_median": int(_stat.median(recs))         if recs else None,
        "concurrent_max":    max(concurs)                if concurs else None,
        "concurrent_median": int(_stat.median(concurs))  if concurs else None,
    }


def _fmt_sec(s: float) -> str:
    """Format seconds to human-readable: 276 → '4m 36s', 14 → '14s'."""
    if s < 60:
        return f"{s:.0f}s"
    m = int(s // 60)
    sec = int(s % 60)
    return f"{m}m {sec}s" if sec else f"{m}m"


def _build_evidence_sentences(rows: list[BenchmarkRow]) -> list[str]:
    """Auto-generate audit-ready evidence sentences for Band C."""
    sentences = []
    for r in rows:
        name    = r.transaction
        cur_fmt = _fmt_sec(r.current_sec)
        action  = r.action or ""
        recs    = f"{r.records:,}" if r.records else None
        concur  = r.concurrent_users

        # Suffix context
        ctx_parts = []
        if recs:
            ctx_parts.append(f"{recs} records")
        if concur:
            ctx_parts.append(f"{concur} concurrent user{'s' if concur != 1 else ''}")
        ctx = f" ({', '.join(ctx_parts)})" if ctx_parts else ""

        if r.baseline_sec > 0:
            base_fmt = _fmt_sec(r.baseline_sec)
            sign     = "better" if r.delta_pct < 0 else "slower"
            delta_abs = abs(round(r.delta_pct, 1))
            if r.status == "BREACH":
                sla_note = f" — SLA BREACH ({_fmt_sec(r.sla_sec)} limit)" if r.sla_sec else " — BREACH threshold exceeded"
                s = f'"{name} {action} completes in {cur_fmt}{ctx} (baseline {base_fmt}) — {delta_abs}% {sign}{sla_note}."'
            elif r.status == "WATCH":
                s = f'"{name} {action} completes in {cur_fmt}{ctx} (baseline {base_fmt}) — {delta_abs}% {sign}, within watch band."'
            else:
                s = f'"{name} {action} completes in {cur_fmt}{ctx} (baseline {base_fmt}) — {delta_abs}% {sign} than benchmark."'
        else:
            sla_note = f" — SLA limit {_fmt_sec(r.sla_sec)}" if r.sla_sec else ""
            s = f'"{name} {action} completes in {cur_fmt}{ctx}{sla_note}. No PROD baseline; captured for reference."'
        sentences.append(s)

    # Sort: BREACH first, WATCH second, then alphabetically
    order = {"BREACH": 0, "WATCH": 1, "OK": 2, "N/A": 3}
    sentences_with_rows = list(zip(sentences, rows))
    sentences_with_rows.sort(key=lambda x: (order.get(x[1].status, 9), x[1].transaction))
    return [s for s, _ in sentences_with_rows]


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.post("/benchmark", response_model=BenchmarkResponse)
async def benchmark_upload(
    file:      UploadFile = File(...),
    threshold: float      = Form(0.0),
) -> BenchmarkResponse:
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Empty file")

    thresh = threshold if threshold > 0 else float(
        config_store.get("benchmark_threshold", 10.0))

    try:
        parsed = _parse_benchmark_file(raw, file.filename or "")
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Cannot parse benchmark file: {exc}") from exc

    rows         = parsed.get("rows", [])
    fill_rate    = parsed.get("fill_rate")
    observations = parsed.get("observations")
    bp_all_rows  = parsed.get("batch_perf_all_rows")
    debug_info   = parsed.get("debug_info", "")

    if not rows and not fill_rate:
        raise HTTPException(status_code=422, detail=(
            f"Could not identify any benchmark data in '{file.filename}'. "
            f"{debug_info} "
            f"Supported formats: (1) PROD + TEST/UAT column pair in any sheet, "
            f"(2) Separate PROD/TEST sheets joined by job name, "
            f"(3) Single TimeTaken column (0H 4M 36S), "
            f"(4) UI Perf sheets with loading/export time columns, "
            f"(5) Flat: Transaction | Baseline (sec) | Current (sec)."
        ).strip())

    # Build batch perf summary before _compute_benchmark (which processes all rows)
    batch_perf_summary = None
    if bp_all_rows is not None:
        batch_perf_summary = _build_batch_perf_summary(bp_all_rows, thresh)

    result_rows, summary = _compute_benchmark(rows, thresh)
    categories = _build_categories(result_rows)

    breaches  = sum(1 for r in result_rows if r.status == "BREACH")
    watches   = sum(1 for r in result_rows if r.status == "WATCH")
    oks       = sum(1 for r in result_rows if r.status == "OK")
    unchanged = sum(1 for r in result_rows if abs(r.delta_pct) <= 1.0)
    sla_brs   = sum(1 for r in result_rows if r.sla_breach)
    avg_delta = round(sum(r.delta_pct for r in result_rows) / len(result_rows), 2) if result_rows else 0.0

    # Single-time capture (no PROD baseline)
    is_single_time = result_rows and all(r.baseline_sec == 0.0 for r in result_rows)
    if is_single_time:
        total = len(result_rows)
        summary = (
            f"📋 UAT performance capture — {total} transaction(s) recorded. "
            f"No PROD baseline loaded; times shown as-is for review. "
            f"Upload a file with PROD + TEST/UAT columns to enable regression comparison."
        )

    # For batch perf files override summary
    elif batch_perf_summary is not None:
        bp = batch_perf_summary
        net = bp["net_delta_secs"]
        direction = "saved" if net >= 0 else "added"
        summary = (
            f"{'⚠️' if bp['regressions'] > 0 else '✅'} "
            f"{bp['regressions']} regression(s), {bp['improvements']} improvement(s) "
            f"across {bp['total_jobs']} jobs. "
            f"Net: {abs(net):.0f}s {direction} per run."
        )

    evidence_sentences = _build_evidence_sentences(result_rows)
    coverage_summary   = _build_coverage_summary(result_rows)

    resp = BenchmarkResponse(
        filename=file.filename or "",
        total_transactions=len(result_rows),
        degraded=breaches, improved=oks, unchanged=unchanged,
        sla_breaches=sla_brs, avg_delta_pct=avg_delta,
        threshold_pct=thresh, rows=result_rows, summary=summary,
        categories=categories,
        fill_rate=fill_rate,
        observations=observations,
        batch_perf_summary=batch_perf_summary,
        evidence_sentences=evidence_sentences,
        coverage_summary=coverage_summary,
    )
    # Persist for narrative/findings server-side fallback + page-reload restore
    try:
        from services import session_cache
        session_cache.set("last_benchmark", resp.model_dump())
    except Exception:
        pass
    try:
        from services.ai_narrator import narrate
        text, model = narrate("benchmark", {
            "summary":      summary,
            "counts":       {"total": len(result_rows), "degraded": breaches,
                             "improved": oks, "unchanged": unchanged,
                             "sla_breaches": sla_brs},
            "avg_delta_pct": avg_delta,
            "threshold_pct": thresh,
            "top_rows":     [r.model_dump() for r in result_rows[:8]],
        })
        if text:
            resp.ai_narrative = text
            resp.ai_model     = model
    except Exception:
        pass
    return resp


@router.post("/benchmark/json", response_model=BenchmarkResponse)
def benchmark_json(body: JsonBenchmarkRequest) -> BenchmarkResponse:
    thresh = float(body.threshold or config_store.get("benchmark_threshold", 10.0))
    result_rows, summary = _compute_benchmark(body.rows or [], thresh)
    categories = _build_categories(result_rows)

    breaches  = sum(1 for r in result_rows if r.status == "BREACH")
    oks       = sum(1 for r in result_rows if r.status == "OK")
    unchanged = sum(1 for r in result_rows if abs(r.delta_pct) <= 1.0)
    sla_brs   = sum(1 for r in result_rows if r.sla_breach)
    avg_delta = round(sum(r.delta_pct for r in result_rows) / len(result_rows), 2) if result_rows else 0.0

    resp = BenchmarkResponse(
        filename=body.filename or "",
        total_transactions=len(result_rows),
        degraded=breaches, improved=oks, unchanged=unchanged,
        sla_breaches=sla_brs, avg_delta_pct=avg_delta,
        threshold_pct=thresh, rows=result_rows, summary=summary,
        categories=categories,
        evidence_sentences=_build_evidence_sentences(result_rows),
        coverage_summary=_build_coverage_summary(result_rows),
    )
    # Persist for narrative/findings server-side fallback + page-reload restore
    try:
        from services import session_cache
        session_cache.set("last_benchmark", resp.model_dump())
    except Exception:
        pass
    try:
        from services.ai_narrator import narrate
        text, model = narrate("benchmark", {
            "summary":      summary,
            "counts":       {"total": len(result_rows), "degraded": breaches,
                             "improved": oks, "unchanged": unchanged,
                             "sla_breaches": sla_brs},
            "avg_delta_pct": avg_delta,
            "threshold_pct": thresh,
            "top_rows":     [r.model_dump() for r in result_rows[:8]],
        })
        if text:
            resp.ai_narrative = text
            resp.ai_model     = model
    except Exception:
        pass
    return resp
