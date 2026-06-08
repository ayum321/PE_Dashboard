"""Quick analysis of Michelin PE files."""
import pandas as pd

csv = r"c:\Users\1039081\Downloads\Michilen_PE\PE signoff\Last_15_Days_Report_of_CS_MICHELIN_SCPO_FF_2025_PROD.csv"
df = pd.read_csv(csv)
df["Start_Time"] = pd.to_datetime(df["Start_Time"])
df["End_Time"] = pd.to_datetime(df["End_Time"])
rc = "Run Time (Sec.) "

print("=" * 60)
print("MICHELIN BATCH DATA ANALYSIS")
print("=" * 60)
print(f"Date Range: {df.Start_Time.min()} → {df.Start_Time.max()}")
print(f"Days span: {(df.Start_Time.max() - df.Start_Time.min()).days}, Unique dates: {df.Start_Time.dt.date.nunique()}")
print(f"Total runs: {len(df)}, Unique jobs: {df['Job_Name'].nunique()}")
print()

print("COMPLETION STATUS:")
for status, count in df["Completion Status"].value_counts().items():
    print(f"  {status}: {count}")

fails = df[df["Completion Status"] == "ENDED NOT OK"]
if len(fails):
    print(f"\nFAILED JOBS ({len(fails)}):")
    for _, r in fails.iterrows():
        print(f"  {r['Job_Name']}: {r['Start_Time']} - {r['Completion Status']}")

print()
rs = df[rc].astype(float)
print(f"Runtime: min={rs.min():.0f}s  max={rs.max():.0f}s ({rs.max()/3600:.3f}h)  mean={rs.mean():.0f}s  median={rs.median():.0f}s")

print("\nTOP 10 LONGEST JOBS:")
top = df.nlargest(10, rc)
for _, r in top.iterrows():
    nm = r["Job_Name"]
    hrs = float(r[rc]) / 3600
    st = r["Completion Status"]
    sa = r["Sub_Application"]
    print(f"  {nm:45s} {hrs:.3f}h  {st:15s}  {sa}")

print("\nSUB_APPLICATION BREAKDOWN:")
for sa in df["Sub_Application"].unique():
    sub = df[df["Sub_Application"] == sa]
    mx = sub[rc].astype(float).max()
    fails_n = (sub["Completion Status"] == "ENDED NOT OK").sum()
    dates = sub.Start_Time.dt.date.nunique()
    print(f"  {sa}: {len(sub)} runs, {dates} days, max={mx/3600:.3f}h, fails={fails_n}")

# SLA XLSX analysis
print("\n" + "=" * 60)
print("SLA XLSX ANALYSIS")
print("=" * 60)
xlsx = r"c:\Users\1039081\Downloads\Michilen_PE\PE signoff\BatchSLA_info.xlsx"
try:
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        ws = wb[sn]
        print(f"\n  Sheet '{sn}': {ws.max_row} rows x {ws.max_column} cols")
        # Print header row
        headers = [str(ws.cell(1, c).value) for c in range(1, min(ws.max_column + 1, 15))]
        print(f"  Headers: {headers}")
        # Print first 3 data rows
        for row in range(2, min(ws.max_row + 1, 5)):
            vals = [str(ws.cell(row, c).value)[:30] for c in range(1, min(ws.max_column + 1, 10))]
            print(f"    Row {row}: {vals}")
except Exception as e:
    print(f"Error: {e}")
